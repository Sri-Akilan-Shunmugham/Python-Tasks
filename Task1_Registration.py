import re
from openpyxl import load_workbook
from goto_py import goto

# getting the Workbook
workbook = load_workbook(filename="login.xlsx")
current_sheet = workbook['Credentials']

# get max column count
max_column = current_sheet.max_column
max_row = current_sheet.max_row
rows = current_sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_column)

print("Welcome User. Select the options below to proceed")
option = input("For Registration press R \nFor Login press L\n")


# Function to check valid username
def checkUsername(usrname):
    regex = re.compile('^[a-zA-Z]+\d*[\._]?[a-zA-Z\d]+@+[a-zA-Z]+\.[a-z]{2,3}$')
    if re.fullmatch(regex, usrname):
        for a, b in rows:
            if a.value == usrname:
                print("Username already in use.\n")
                return False
            else:
                continue
        return True

    return False


# Function to check valid password
def checkPwd(pswd):
    regex = re.compile("^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*#?&])[A-Za-z\d@$!#%*?&]{5,16}$")
    if re.fullmatch(regex, pswd):
        return True

    return False


# Checking option for Registration
if option == 'R' or option == 'r':
    print("Username/Mail Should be valid Mail Id)")
    username = input("Please Enter Username:")
    checkUsr = checkUsername(username)

    if checkUsr:
        print("\nPassword should have minimum \n1) One Upper Case Character\n2) One Lower Case Character\n3) One Digit "
              "\n4) One Special Character\n5) Length Should be 8-18\n")
        pwd = input("Please Enter Password:")
        checkPass = checkPwd(pwd)

        if checkPass:
            cell1 = current_sheet.cell(max_row + 1, 1)
            cell1.value = username
            cell2 = current_sheet.cell(max_row + 1, 2)
            cell2.value = pwd
            print("Registration Successful")


        else:
            print("Password Invalid. \n1) Press X to Close.\n2) Press R to Restart")
            T = input()
            if T == 'R' or T == 'r':
                username = input("Please Enter Username:")
                checkUsr = checkUsername(username)

                if checkUsr:
                    print(
                        "\nPassword should have minimum \n1) One Upper Case Character\n2) One Lower Case "
                        "Character\n3) One Digit \n4) One Special Character\n5) Length Should be 8-18\n")
                    pwd = input("Please Enter Password:")
                    checkPass = checkPwd(pwd)

                    if checkPass:
                        cell1 = current_sheet.cell(max_row + 1, 1)
                        cell1.value = username
                        cell2 = current_sheet.cell(max_row + 1, 2)
                        cell2.value = pwd
                        print("Registration Successful")
                        workbook.save(filename="login.xlsx")
                        workbook.close()
                        exit(0)
                    else:
                        print("Invalid Password. Try again Next time")
                        workbook.save(filename="login.xlsx")
                        workbook.close()
                        exit(0)
                else:
                    print("Invalid Username, Try again Next time")
                    workbook.save(filename="login.xlsx")
                    workbook.close()
                    exit(0)
            elif T == 'X' or T == 'x':
                print("Have a nice day. Bye")
                workbook.save(filename="login.xlsx")
                workbook.close()
                exit(0)
    else:
        print("Invalid User name/mail Id. \n1) Press X to Close.\n2) Press R to Restart")
        T = input()
        if T == 'R' or T == 'r':
            username = input("Please Enter Username:")
            checkUsr = checkUsername(username)

            if checkUsr:
                print(
                    "\nPassword should have minimum \n1) One Upper Case Character\n2) One Lower Case "
                    "Character\n3) One Digit \n4) One Special Character\n5) Length Should be 8-18\n")
                pwd = input("Please Enter Password:")
                checkPass = checkPwd(pwd)

                if checkPass:
                    cell1 = current_sheet.cell(max_row + 1, 1)
                    cell1.value = username
                    cell2 = current_sheet.cell(max_row + 1, 2)
                    cell2.value = pwd
                    print("Registration Successful")
                    workbook.save(filename="login.xlsx")
                    workbook.close()
                    exit(0)
                else:
                    print("Invalid Password. Try again Next time")
                    workbook.save(filename="login.xlsx")
                    workbook.close()
                    exit(0)
            else:
                print("Invalid Username, Try again Next time")
                workbook.save(filename="login.xlsx")
                workbook.close()
                exit(0)
        elif T == 'X' or T == 'x':
            print("Have a nice day. Bye")
            workbook.save(filename="login.xlsx")
            workbook.close()
            exit(0)

# Checking option for Login
elif option == 'L' or option == 'l':
    print("Start login")
    username = input("Enter Username: ")
    for a, b in rows:
        if a.value == username:

            pwd = input("Enter Password: ")
            if b.value == pwd:
                print("Login Successful")
                workbook.save(filename="login.xlsx")
                workbook.close()
                exit(0)
            else:
                print("Password does not match Username")
                pwd = input("Enter Password:")
                if b.value == pwd:
                    print("login Successful.")
                    workbook.save(filename="login.xlsx")
                    workbook.close()
                    exit(0)
                else:
                    print("Password does not match Username")
                    T = input("Press F for Forget Password.\nPress N to set new Password\n")
                    if T == 'F' or T == 'f':
                        print("Registered password for "+a.value+" is "+b.value)
                        workbook.save(filename="login.xlsx")
                        workbook.close()
                        exit(0)
                    elif T == 'N' or T == 'n':
                        pwd = input("Enter New Password: ")
                        b.value = pwd
                        print("Password Updated")
                        workbook.save(filename="login.xlsx")
                        workbook.close()
                        exit(0)
                    else:
                        print("You chose "+T+". Bye")
                        workbook.save(filename="login.xlsx")
                        workbook.close()
                        exit(0)
        else:
            continue

    for a, b in rows:
        if a.value != username:
            print("User name not available.")
            T = input("Do you like to Register (Y/N): ")
            if T == 'Y' or T == 'y':
                username = input("Please Enter Username:")
                checkUsr = checkUsername(username)

                if checkUsr:
                    print(
                        "\nPassword should have minimum \n1) One Upper Case Character\n2) One Lower Case "
                        "Character\n3) One Digit \n4) One Special Character\n5) Length Should be 8-18\n")
                    pwd = input("Please Enter Password:")
                    checkPass = checkPwd(pwd)

                    if checkPass:
                        cell1 = current_sheet.cell(max_row + 1, 1)
                        cell1.value = username
                        cell2 = current_sheet.cell(max_row + 1, 2)
                        cell2.value = pwd
                        print("Registration Successful")
                        workbook.save(filename="login.xlsx")
                        workbook.close()
                        exit(0)
                    else:
                        print("Invalid Password. Try again Next time")
                        workbook.save(filename="login.xlsx")
                        workbook.close()
                        exit(0)
                else:
                    print("Invalid Username, Try again Next time")
                    workbook.save(filename="login.xlsx")
                    workbook.close()
                    exit(0)

            elif T == 'N' or T == 'n':
                print("Have a nice day. Bye")
                workbook.save(filename="login.xlsx")
                workbook.close()
                exit(0)
else:
    print("You chose "+option+". Rerun to start the process again")

workbook.save(filename="login.xlsx")
workbook.close()
